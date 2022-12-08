import glob, os, sys
from os.path import join
import re
import time
from datetime import date, datetime
from time import strptime
from dateutil import parser
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import shutil
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import pandas as pd
import os

fdname = os.path.dirname(__file__)+'/'

# fdname="E:/Raffy/datascrap/"
#fname="vest_Women_Accessories.xlsx"
fname="All_files/test_vest3.xlsx"
df = pd.read_excel(fname)
#print(df)



def log1(gender1,brand1,category1,subcategory,proddesc,prodcon,material1,colour1,price1,website,location1,cdate,purl):
    try:
        today = datetime.today().strftime('%d/%m/%Y')
        if not os.path.exists(fdname + 'log.xlsx'):
            book = Workbook()
            ws = book.active
            
            ws.cell(1,1).value = 'Gender'
            ws.cell(1,2).value = 'Brand'
            ws.cell(1,3).value = 'Category'
            ws.cell(1,4).value = 'Sub Category'
            ws.cell(1,5).value = 'Product Desc'
            ws.cell(1,6).value = 'Product Condition'
            ws.cell(1,7).value = 'Material'
            ws.cell(1,8).value = 'Colour'
            ws.cell(1,9).value = 'Price'
            ws.cell(1,10).value = 'Website'            
            ws.cell(1,11).value = 'Location'
            ws.cell(1,12).value = 'Datetimestamp'
            ws.cell(1,13).value = 'purl'
            
            
            
            
            ws.column_dimensions["A"].width = 30.0
            ws.column_dimensions["B"].width = 30.0
            ws.column_dimensions["C"].width = 30.0
            ws.column_dimensions["D"].width = 30.0
            ws.column_dimensions["E"].width = 30.0
            ws.column_dimensions["F"].width = 30.0
            ws.column_dimensions["G"].width = 30.0
            ws.column_dimensions["H"].width = 30.0
            ws.column_dimensions["I"].width = 30.0
            ws.column_dimensions["J"].width = 30.0
            ws.column_dimensions["K"].width = 30.0
            ws.column_dimensions["L"].width = 30.0
            ws.column_dimensions["M"].width = 30.0
            ws.column_dimensions["N"].width = 30.0
            
            
            
            
            
            ws['A1'].font = Font(bold=True)
            ws['B1'].font = Font(bold=True)
            ws['C1'].font = Font(bold=True)
            ws['D1'].font = Font(bold=True)
            ws['E1'].font = Font(bold=True)
            ws['F1'].font = Font(bold=True)
            ws['G1'].font = Font(bold=True)
            ws['H1'].font = Font(bold=True)
            ws['I1'].font = Font(bold=True)
            ws['J1'].font = Font(bold=True)
            ws['K1'].font = Font(bold=True)
            ws['L1'].font = Font(bold=True)
            ws['M1'].font = Font(bold=True)
            ws['N1'].font = Font(bold=True)
            
            
            
            
            book.save(fdname +'log.xlsx')
            
        book2 = load_workbook(fdname +'log.xlsx')
        sheet = book2.active
                
        sheet.append((gender1,brand1,category1,subcategory,proddesc,prodcon,material1,colour1,price1,website,location1,cdate,purl))
        sheet.column_dimensions["A"].width = 30.0
        sheet.column_dimensions["B"].width = 30.0
        sheet.column_dimensions["C"].width = 30.0
        sheet.column_dimensions["D"].width = 30.0
        sheet.column_dimensions["E"].width = 30.0
        sheet.column_dimensions["F"].width = 30.0
        sheet.column_dimensions["G"].width = 30.0
        sheet.column_dimensions["H"].width = 30.0
        sheet.column_dimensions["I"].width = 30.0
        sheet.column_dimensions["J"].width = 30.0
        sheet.column_dimensions["K"].width = 30.0
        sheet.column_dimensions["L"].width = 30.0
        sheet.column_dimensions["M"].width = 30.0
        sheet.column_dimensions["N"].width = 30.0
        
        
        
        book2.save(fdname +'log.xlsx')
    except Exception as e:
        print(e)
        pass
        
def log2(purl):
    try:
        
        if not os.path.exists(fdname + 'errorlog.xlsx'):
            book = Workbook()
            ws = book.active
            
            ws.cell(1,1).value = 'Url'
            ws.column_dimensions["A"].width = 30.0
            ws['A1'].font = Font(bold=True)
            
            book.save(fdname +'errorlog.xlsx')
            
        book2 = load_workbook(fdname +'errorlog.xlsx')
        sheet = book2.active
                
        sheet.append([purl])
        sheet.column_dimensions["A"].width = 30.0
        # print(purl)
        book2.save(fdname +'errorlog.xlsx')
    except Exception as e:
        print(e)
        pass

chrome_options = Options()
#chrome_options.add_argument("--headless")
chrome_options.add_argument('--no-sandbox')
chrome_options.add_argument("--disable-dev-shm-usage")
chrome_options.add_argument("--ignore-certificate-error")
chrome_options.add_argument("--ignore-ssl-errors")
#chrome_options.add_experimental_option("debuggerAddress", "localhost:9222")
browser = webdriver.Chrome(f'{fdname}chromedriver.exe', options=chrome_options)
#print(browser.capabilities)
browser.set_page_load_timeout(20)

i=0
try:
    for ind in df.index:
        purl=str(df['prod_url'][ind].strip())               
        try:
            if("undefined" not in df['prod_url'][ind]):
                browser.get(df['prod_url'][ind].strip())
                
                if(i==0):
                    browser.maximize_window()
                    time.sleep(3)
                    # allow cookies
                    browser.find_element("xpath", '//*[@id="popin_tc_privacy_button_2"]').click()
                    
                    time.sleep(3)  
                    try:
                        browser.find_element("xpath", '//*[@id="cross-x-thick"]').click()
                    except:
                        print('Registration page closed')
                    
                    
                    time.sleep(10)  
                    try:
                        # clicks footer
                        browser.find_element(By.XPATH, '//*[@id="footer"]/div[1]/div/div[2]/button').click()
                        
                        time.sleep(3)  
                        # clicks Currency section
                        browser.find_element("name","currency").click()

                        # clicks the GPB
                        try:
                            browser.find_element(By.XPATH, '/html/body/div[16]/div/div/div/div/div/form/div[1]/div[3]/div/select/option[3]').click()
                        except:
                            browser.find_element(By.XPATH, '/html/body/div[15]/div/div/div/div/div/form/div[1]/div[3]/div/select/option[3]').click()
                            
                        # clicks the SAVE CHANGES
                        try:
                            browser.find_element(By.XPATH, '/html/body/div[16]/div/div/div/div/div/form/div[2]/div/button').click()
                        except:
                            browser.find_element(By.XPATH, '/html/body/div[15]/div/div/div/div/div/form/div[2]/div/button').click()
                        
                        print('preference updated..') 

                    except:
                        print('Error updating preference..')   


                price=browser.find_element('xpath',"//*[@id='__next']/div/main/div/div[4]/div/div[1]/div/div[2]/div")
                price1=price.text
                browser.find_element('xpath',"//*[@id='__next']/div/main/section[1]/div[1]/div/div[2]/div[1]/div/button[1]/span").click()
                time.sleep(1)
                
                try:
                    desc1=browser.find_element('xpath','_by_xpath',"//*[@id='__next']/div/main/section[1]/div[1]/div/div[2]/div[1]/p[1]")
                except:
                    desc1=browser.find_element('xpath',"//*[@id='__next']/div/main/section[1]/div[1]/div/div[2]/div[1]/p")
                    pass
                
                proddesc=desc1.text
                
                prod_details = browser.find_element("xpath", '//*[@id="__next"]/div/main/section[1]/div[1]/div/div[2]/div[2]').text
                
                gender1 = (re.search("[Cc]ategories {0,1}:{0,1} {0,1}(.*)\n",prod_details).group(1)) 
                colour1 = (re.search("[Cc]olou{0,1}r {0,1}:{0,1} {0,1}(.*)\n",prod_details).group(1)) 
                material1 = (re.search("[Mm]aterial {0,1}:{0,1} {0,1}(.*)\n",prod_details).group(1)) 
                location1 = (re.search("[Ll]ocation {0,1}:{0,1} {0,1}(.*)\n",prod_details).group(1)) 
                brand1 = (re.search("[Dd]esigner {0,1}:{0,1} {0,1}(.*)\n",prod_details).group(1)) 
                subcategory = (re.search("[Ss]ub-category {0,1}:{0,1} {0,1}(.*)\n",prod_details).group(1)) 
                category1 = (re.search("[Cc]ategory {0,1}:{0,1} {0,1}(.*)\n",prod_details).group(1)) 
                prodcon = (re.search("[Cc]ondition {0,1}:{0,1} {0,1}(.*)\n",prod_details).group(1)).replace("More info","")
                
                    
                cdate=datetime.now()
                log1(gender1,brand1,category1,subcategory,proddesc,prodcon,material1,colour1,price1,"Vestiaire",location1,cdate,purl)
                print(i)    
                i=i+1
                #time.sleep(3)
        except Exception as e :
            log2(str(purl))
            print('Exception arised..')
            print(i)    
            i=i+1
            pass
            
    
except Exception as e:
    print(e)
    exc_type, exc_obj, exc_tb = sys.exc_info()
    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
    print(exc_type, fname, exc_tb.tb_lineno)
    print('OOPS!, Something went wrong.')
    
os.startfile('alarm.mp3')

