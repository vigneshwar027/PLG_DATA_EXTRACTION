import glob, os, sys
from os.path import join
import re
import time
from datetime import date, datetime
from time import strptime
from dateutil import parser
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import shutil
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import pandas as pd

fdname = os.path.dirname(__file__)+'/'
cdate=datetime.now()

chrome_options = Options()
# chrome_options.add_experimental_option("debuggerAddress","localhost:8989")

# chrome_options.add_argument("--headless")
chrome_options.add_argument('--no-sandbox')
chrome_options.add_argument("--disable-dev-shm-usage")
#chrome_options.add_experimental_option("debuggerAddress", "localhost:9222")
browser = webdriver.Chrome(fdname+'chromedriver.exe', options=chrome_options)
# url = browser.command_executor._url       #"http://127.0.0.1:60622/hub"
# session_id = browser.session_id 

# browser = webdriver.Remote(command_executor=url,desired_capabilities={})
# browser.close()   # this prevents the dummy browser
# browser.session_id = session_id

#print(browser.capabilities)
i=3


def extract_data_vest(df):
    continous_exp_count = 0
    OP_df = pd.DataFrame({'Gender':[],"Description":[],	"Price":[],	"Location":[],	"Color":[],	"Material":[],	"Condition":[],	"Designer":[],	"SubCat":[],	"Category":[],	"WebSite":[],	"TimeStamp":[],	"product_url":[]})

    err_df = pd.DataFrame({"error_url":[],"Exception":[]})

    for index, url in enumerate(df['prod_url']): 

        try:
            browser.get(url)
            # time.sleep(3)
            
            if index == 0:
                time.sleep(3)
                browser.maximize_window()
                # allow cookies
                browser.find_element("xpath", '//*[@id="popin_tc_privacy_button_2"]').click()
                
                time.sleep(3)  
                try:
                    browser.find_element("xpath", '//*[@id="cross-x-thick"]').click()
                except:
                    print('Registration page closed')   
                
            if True:
                
                price=browser.find_element_by_xpath("//*[@id='__next']/div/main/div/div[4]/div/div[1]/div/div[2]/div").text
                
                # click read more button if exist
                try:
                    browser.find_element("xpath", '//*[@id="__next"]/div/main/section[1]/div[1]/div/div[2]/div[1]/div/button[1]').click()
                except:
                    pass
                
                prod_desc = browser.find_element("xpath", '//*[@id="__next"]/div/main/section[1]/div[1]/div/div[2]/div[1]/p[1]').text
                
                prod_details = browser.find_element("xpath", '//*[@id="__next"]/div/main/section[1]/div[1]/div/div[2]/div[2]').text
                
                gender = (re.search("[Cc]ategories {0,1}:{0,1} {0,1}(.*)\n",prod_details).group(1)) 
                color = (re.search("[Cc]olou{0,1}r {0,1}:{0,1} {0,1}(.*)\n",prod_details).group(1)) 
                material = (re.search("[Mm]aterial {0,1}:{0,1} {0,1}(.*)\n",prod_details).group(1)) 
                location = (re.search("[Ll]ocation {0,1}:{0,1} {0,1}(.*)\n",prod_details).group(1)) 
                brand = (re.search("[Dd]esigner {0,1}:{0,1} {0,1}(.*)\n",prod_details).group(1)) 
                sub_cat = (re.search("[Ss]ub-category {0,1}:{0,1} {0,1}(.*)\n",prod_details).group(1)) 
                category = (re.search("[Cc]ategory {0,1}:{0,1} {0,1}(.*)\n",prod_details).group(1)) 
                condition = (re.search("[Cc]ondition {0,1}:{0,1} {0,1}(.*)\n",prod_details).group(1))


                current_df = pd.DataFrame({'Gender':[gender],"Description":[prod_desc],	"Price":[price],"Location":[location],"Color":[color],"Material":[material],"Condition":[condition],"Designer":[brand],"SubCat":[sub_cat],"Category":[category],"WebSite":['Vestiare'],"TimeStamp":[str(datetime.now())],"product_url":[url]})

                OP_df = pd.concat([OP_df,current_df])
                
                # resetting continous exception count
                continous_exp_count = 0
                # quit()

        except Exception as excep:
            continous_exp_count+= 1
            current_err =  pd.DataFrame({"error_url":[url],"Exception":[str(excep)]})
            err_df = pd.concat([err_df,current_err])

            if continous_exp_count >= 10:
                break #the loop
        print(continous_exp_count)
    rows = OP_df.values.tolist()
    workbook = load_workbook(filename="All_files/processed_file.xlsx")
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    workbook.save(filename="All_files/processed_file.xlsx")
    
    rows = err_df.values.tolist()
    workbook = load_workbook(filename="All_files/errors_file.xlsx")
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    workbook.save(filename="All_files/errors_file.xlsx")

    # OP_df.to_excel("processed_file.xlsx")
    # err_df.to_excel("All_files/errors_file.xlsx")
    browser.close()



#############################



def extract_data_lampoo(df):

    continous_exp_count = 0

    OP_df = pd.DataFrame({'Gender':[],"Description":[],	"Price":[],	"Location":[],	"Color":[],	"Material":[],	"Condition":[],	"Designer":[],	"SubCat":[],"Category":[],"Status":[],"WebSite":[],	"TimeStamp":[],"product_url":[]})

    err_df = pd.DataFrame({"error_url":[],'Gender':[],"Category":[],"SubCat":[],"Exception":[],"TimeStamp":[]})

    # for index, url in enumerate(df['prod_url']): 
    for index,rw in df.iterrows():
        print(index)
        try:
            browser.get(rw['prod_url'])
            
            if index == 0:
                time.sleep(5)
                browser.maximize_window()
                # allow cookies
                browser.find_element("xpath", '/html/body/div[2]/div[2]/div/div[1]/div/div[2]/div/button[2]').click()
                
            if True:
                # time.sleep(5)
                
                gender = rw['gender']
                location = 'not found'
                sub_cat = rw['sub_category']
                category = rw['category']   
                
                price = browser.find_element(By.CLASS_NAME, "product-info-price").text

                # material = browser.find_element(By.CLASS_NAME, "additional-attributes attributes-composition").text
                
                # color = browser.find_element(By.CLASS_NAME, "additional-attributes attributes-color").text
                
                try:
                    color = browser.find_element("xpath", '//*[@id="product-attribute-specs-table"]/div[2]/ul').text
                except:
                    try:
                        color = browser.find_element("xpath", '//*[@id="product-attribute-specs-table"]/div[1]/ul').text
                    except:
                        color = 'not found'

                try:
                    material = browser.find_element("xpath", '//*[@id="product-attribute-specs-table"]/div[3]/dl/div/dd/ul').text
                except:
                    try:
                        material = browser.find_element("xpath", '//*[@id="product-attribute-specs-table"]/div[2]/dl/div/dd/ul').text
                    except:
                        try:
                            material = browser.find_element("xpath", '//*[@id="product-attribute-specs-table"]/div[4]/dl/div/dd/ul').text
                        except:
                            try:
                                material = browser.find_element("xpath", '//*[@id="product-attribute-specs-table"]/div/dl/div/dd/ul').text
                            except:
                                try:
                                    material = browser.find_element("xpath", '//*[@id="product-attribute-specs-table"]/div[3]/div').text
                                except:
                                    try:
                                        material = browser.find_element("xpath", '//*[@id="product-attribute-specs-table"]/div[4]/div').text
                                    except:
                                        material = 'not found'

                brand = browser.find_element("xpath", '//*[@id="maincontent"]/div[3]/div/div[1]/h1/span[1]').text

                condition = browser.find_element("xpath", '//*[@id="data-condition"]/div/div').text

                prod_desc = browser.find_element("xpath", '//*[@id="data-all_description"]/div/div').text

                try:
                    pstatus = browser.find_element("xpath", '//*[@id="product-addtocart-button"]').text
                except:
                    try:
                        pstatus = browser.find_element("xpath", '//*[@id="product_addtocart_form"]/div/div/div/button').text  
                    except:
                        pstatus = browser.find_element("xpath", '//*[@id="maincontent"]/div[3]/div/div[1]/div[1]/div/div/button').text  

                current_df = pd.DataFrame({'Gender':[gender],"Description":[prod_desc],	"Price":[price],"Location":[location],"Color":[color],"Material":[material],"Condition":[condition],"Designer":[brand],"SubCat":[sub_cat],"Category":[category],"Status":[pstatus],"WebSite":['Lampoo'],"TimeStamp":[str(datetime.now())],"product_url":[rw['prod_url']]})

                OP_df = pd.concat([OP_df,current_df])
                
                # resetting continous exception count
                continous_exp_count = 0
                # quit()

        except Exception as excep:
            continous_exp_count+= 1
            print('Exception arised..')
            current_err =  pd.DataFrame({"error_url":[rw['prod_url']],'Gender':[rw['gender']],"Category":[rw['category']],"SubCat":[rw['sub_category']],"Exception":[str(excep)],"TimeStamp":[str(datetime.now())]})
            
            err_df = pd.concat([err_df,current_err])

            if continous_exp_count >= 15:
                break #the loop

        # print(continous_exp_count)
    rows = OP_df.values.tolist()
    workbook = load_workbook(filename="All_files/lampoo_processed_file.xlsx")
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    workbook.save(filename="All_files/lampoo_processed_file.xlsx")
    
    rows = err_df.values.tolist()
    workbook = load_workbook(filename="All_files/lampoo_errors_file.xlsx")
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    workbook.save(filename="All_files/lampoo_errors_file.xlsx")

    # OP_df.to_excel("processed_file.xlsx")
    # err_df.to_excel("All_files/errors_file.xlsx")
    browser.close()


def start():
    df_vest = pd.read_excel(r"All_files/source_url_vest.xlsx")
    extract_data_vest(df_vest)
    
    # df_lampoo = pd.read_excel(r"All_files/test.xlsx")
    # extract_data_lampoo(df_lampoo)

start()
os.startfile('alarm.mp3')
print('Extraction Complete..') 