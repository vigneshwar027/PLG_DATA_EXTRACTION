import glob, os, sys
from os.path import join
import re
import time
from datetime import date, datetime
from time import strptime
from dateutil import parser
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import shutil
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import pandas as pd


fdname= os.path.dirname(__file__)+'/'
#fname="vest_Women_Accessories.xlsx"
fname="All_files/test.xlsx"
df = pd.read_excel(fname)
#print(df)

def log1(gender1,brand1,category1,subcategory,proddesc,prodcon,material1,colour1,price1,website,location1,cdate,purl,pstatus):
    try:
        today = datetime.today().strftime('%d/%m/%Y')
        if not os.path.exists(fdname + 'log.xlsx'):
            book = Workbook()
            ws = book.active
            
            ws.cell(1,1).value = 'Gender'
            ws.cell(1,2).value = 'Brand'
            ws.cell(1,3).value = 'Category'
            ws.cell(1,4).value = 'Sub Category'
            ws.cell(1,5).value = 'Product Desc'
            ws.cell(1,6).value = 'Product Condition'
            ws.cell(1,7).value = 'Material'
            ws.cell(1,8).value = 'Colour'
            ws.cell(1,9).value = 'Price'
            ws.cell(1,10).value = 'Website'            
            ws.cell(1,11).value = 'Location'
            ws.cell(1,12).value = 'Datetimestamp'
            ws.cell(1,13).value = 'url'
            ws.cell(1,14).value = 'status'
            
            
            
            
            ws.column_dimensions["A"].width = 30.0
            ws.column_dimensions["B"].width = 30.0
            ws.column_dimensions["C"].width = 30.0
            ws.column_dimensions["D"].width = 30.0
            ws.column_dimensions["E"].width = 30.0
            ws.column_dimensions["F"].width = 30.0
            ws.column_dimensions["G"].width = 30.0
            ws.column_dimensions["H"].width = 30.0
            ws.column_dimensions["I"].width = 30.0
            ws.column_dimensions["J"].width = 30.0
            ws.column_dimensions["K"].width = 30.0
            ws.column_dimensions["L"].width = 30.0
            ws.column_dimensions["M"].width = 30.0
            ws.column_dimensions["N"].width = 30.0
            ws.column_dimensions["O"].width = 30.0
            
            
            
            
            
            ws['A1'].font = Font(bold=True)
            ws['B1'].font = Font(bold=True)
            ws['C1'].font = Font(bold=True)
            ws['D1'].font = Font(bold=True)
            ws['E1'].font = Font(bold=True)
            ws['F1'].font = Font(bold=True)
            ws['G1'].font = Font(bold=True)
            ws['H1'].font = Font(bold=True)
            ws['I1'].font = Font(bold=True)
            ws['J1'].font = Font(bold=True)
            ws['K1'].font = Font(bold=True)
            ws['L1'].font = Font(bold=True)
            ws['M1'].font = Font(bold=True)
            ws['N1'].font = Font(bold=True)
            ws['O1'].font = Font(bold=True)
            
            
            
            
            book.save(fdname +'log.xlsx')
            
        book2 = load_workbook(fdname +'log.xlsx')
        sheet = book2.active
                
        sheet.append((gender1,brand1,category1,subcategory,proddesc,prodcon,material1,colour1,price1,website,location1,cdate,purl,pstatus))
        sheet.column_dimensions["A"].width = 30.0
        sheet.column_dimensions["B"].width = 30.0
        sheet.column_dimensions["C"].width = 30.0
        sheet.column_dimensions["D"].width = 30.0
        sheet.column_dimensions["E"].width = 30.0
        sheet.column_dimensions["F"].width = 30.0
        sheet.column_dimensions["G"].width = 30.0
        sheet.column_dimensions["H"].width = 30.0
        sheet.column_dimensions["I"].width = 30.0
        sheet.column_dimensions["J"].width = 30.0
        sheet.column_dimensions["K"].width = 30.0
        sheet.column_dimensions["L"].width = 30.0
        sheet.column_dimensions["M"].width = 30.0
        sheet.column_dimensions["N"].width = 30.0
        sheet.column_dimensions["O"].width = 30.0
        
        
        
        book2.save(fdname +'log.xlsx')
    except Exception as e:
        print(e)
        pass
        
def log2(purl):
    try:
        
        if not os.path.exists(fdname + 'errorlog.xlsx'):
            book = Workbook()
            ws = book.active
            
            ws.cell(1,1).value = 'Url'
            ws.column_dimensions["A"].width = 30.0
            ws['A1'].font = Font(bold=True)
            
            book.save(fdname +'errorlog.xlsx')
            
        book2 = load_workbook(fdname +'errorlog.xlsx')
        sheet = book2.active
                
        sheet.append([purl])
        sheet.column_dimensions["A"].width = 30.0
        # print(purl)
        book2.save(fdname +'errorlog.xlsx')
    except Exception as e:
        print(e)
        pass

chrome_options = Options()
#chrome_options.add_argument("--headless")
chrome_options.add_argument('--no-sandbox')
chrome_options.add_argument("--disable-dev-shm-usage")
chrome_options.add_argument("--ignore-certificate-error")
chrome_options.add_argument("--ignore-ssl-errors")
#chrome_options.add_experimental_option("debuggerAddress", "localhost:9222")
browser = webdriver.Chrome(f'{fdname}chromedriver.exe', options=chrome_options)
#print(browser.capabilities)

i=0
try:
    for ind in df.index:
        purl=str(df['prod_url'][ind].strip())               
        try:
            if("undefined" not in df['prod_url'][ind]):
                browser.get(df['prod_url'][ind].strip())
                
                if(i==0):
                    browser.maximize_window()
                    time.sleep(3)
                    # allow cookies
                    browser.find_element("xpath", '/html/body/div[2]/div[2]/div/div[1]/div/div[2]/div/button[2]').click()
                    
                gender1 = df['gender'][ind].strip()
                location1 = ''
                subcategory = df['sub_category'][ind].strip()
                category1 = df['category'][ind].strip()  
                # print (gender1)
                # price = browser.find_element("xpath", '//*[@id="maincontent"]/div[3]/div/div[1]/div[2]/div[1]').text

                price1 = browser.find_element(By.CLASS_NAME, "product-info-price").text
                
                # material = browser.find_element(By.CLASS_NAME, "additional-attributes attributes-composition").text
                
                # color = browser.find_element(By.CLASS_NAME, "additional-attributes attributes-color").text
                
                try:
                    colour1 = browser.find_element("xpath", '//*[@id="product-attribute-specs-table"]/div[2]/ul').text
                except:
                    try:
                        colour1 = browser.find_element("xpath", '//*[@id="product-attribute-specs-table"]/div[1]/ul').text
                    except:
                        colour1 = ''
                        pass
               
                try:
                    material1 = browser.find_element("xpath", '//*[@id="product-attribute-specs-table"]/div[3]/dl/div/dd/ul').text
                except:
                    try:
                        material1 = browser.find_element("xpath", '//*[@id="product-attribute-specs-table"]/div[2]/dl/div/dd/ul').text
                    except:
                        try:
                            material1 = browser.find_element("xpath", '//*[@id="product-attribute-specs-table"]/div[4]/dl/div/dd/ul').text
                        except:
                            try:
                                material1 = browser.find_element("xpath", '//*[@id="product-attribute-specs-table"]/div/dl/div/dd/ul').text
                            except:
                                try:
                                    material1 = browser.find_element("xpath", '//*[@id="product-attribute-specs-table"]/div[3]/div').text
                                except:
                                    material1 = ''
                                pass
                                

                brand1 = browser.find_element("xpath", '//*[@id="maincontent"]/div[3]/div/div[1]/h1/span[1]').text

                prodcon = browser.find_element("xpath", '//*[@id="data-condition"]/div/div').text

                proddesc = browser.find_element("xpath", '//*[@id="data-all_description"]/div/div').text 
                
                try:
                    pstatus = browser.find_element("xpath", '/html/body/div[1]/main/div[3]/div/div[1]/div[1]/div/div/button').text
                except:
                    try:
                        pstatus = browser.find_element("xpath", '/html/body/div[1]/main/div[3]/div/div[1]/div[3]/div/form/div/div/div/button').text  
                    except:
                        pstatus=""
                        pass                        
                
                                
                cdate=datetime.now()
                log1(gender1,brand1,category1,subcategory,proddesc,prodcon,material1,colour1,price1,"Lampoo",location1,cdate,purl,pstatus)
                print(i)    
                i=i+1
                #time.sleep(3)
        except Exception as e:
            log2(str(purl))
            print(e)
            print(i)    
            i=i+1
            pass
            
        
except Exception as e:
    print(e)
    exc_type, exc_obj, exc_tb = sys.exc_info()
    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
    print(exc_type, fname, exc_tb.tb_lineno)
    print('OOPS!, Something went wrong.')
    

browser.close()
